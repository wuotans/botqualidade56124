import json
import logging
import os
import re
import shutil
import time
from datetime import datetime, date, timedelta
from priority_classes.decorators.decorators import time_out, TimeoutException
import polars as pl
import pandas as pd


class Handler:

    def __init__(self):
        """
        This class do not initiate any self value, its use to handle generic data involving pandas, os, json, datetime
        regular expressions, windows processes and so on.

        """
        pass

    @staticmethod
    def create_folder(path):
        """
        Creates a folder at a given path if it does not exist already.

        :param path: The path where the folder should be created
        :type path: str
        :return: The absolute path of the created folder
        :rtype: str
        """

        if ':' not in path:
            root_path = os.getcwd()
            path = os.path.abspath(root_path + f'/{path}')

        if not os.path.exists(path):
            os.mkdir(path)
        return path

    def create_file_txt(self, text: str = '', name_file: str = '', path_to_save: str = '', subs: bool = False):
        """
        Creates a .txt file at a specified location with provided text. If the file already exists,
        the method can either overwrite it or read its content based on the 'subs' argument.
        The method first ensures that the path to save the file exists, creating it if it doesn't.

        :param text: The text to write into the file. Default is an empty string.
        :type text: str

        :param name_file: The name of the file to be created. Default is an empty string.
        :type name_file: str

        :param path_to_save: The directory where the file should be saved. Default is the current working directory.
        :type path_to_save: str

        :param subs: If True, existing file will be overwritten. If False and file exists, its content will be read. Default is False.
        :type subs: bool

        :return: The text written to the file or read from it.
        :rtype: str

        :Example:

            self._create_file_txt("Hello, world!", "test", "/path/to/save", True) # creates/overwrites a file named "test.txt" with the text "Hello, world!" at "/path/to/save"

        """
        path_to_save = self.create_folder(path_to_save)
        full_path_file = f'{path_to_save}\{name_file}.txt'

        if not os.path.exists(full_path_file) or subs:
            with open(full_path_file, 'w') as f:
                f.write(text)
        else:
            with open(full_path_file, 'r') as f:
                text = f.read()
        return text

    def create_file_json(self, dict_file: dict, name_file: str = '', path_to_save: str = '', subs: bool = False):
        path_to_save = self.create_folder(path_to_save)
        full_path_file = f'{path_to_save}\{name_file}.json'

        if not os.path.exists(full_path_file) or subs:
            with open(full_path_file, 'w') as f:
                json.dump(dict_file, f)
        else:
            with open(full_path_file, "r") as readit:
                dict_file = json.load(readit)

        return dict_file

    @staticmethod
    def get_dummy():
        # Get the current time in seconds since the epoch
        timestamp = time.time()

        # Convert the timestamp to milliseconds
        dummy = int(timestamp * 1000)

        logging.info(dummy)
        return dummy

    @staticmethod
    def move_folder(folder_copy, folder_write):
        """
        Move a folder from the source location to the destination location.

        :param folder_copy: Source folder to be moved.
        :type folder_copy: str
        :param folder_write: Destination folder where the source folder will be moved to.
        :type folder_write: str
        """
        shutil.copytree(folder_copy, folder_write)

    def delete_files_folder(self, folder):
        """
        Delete all files and subfolders in a folder.

        :param folder: Folder path.
        :type folder: str
        :return: True if the operation is successful, False otherwise.
        :rtype: bool
        """
        if os.path.exists(folder):
            shutil.rmtree(folder)
            self.create_folder(folder)
        else:
            self.create_folder(folder)
        return True

    @staticmethod
    def delete_file(path_file):
        """
        Delete a file at the specified path.

        :param path_file: File path.
        :type path_file: str
        """
        try:
            if os.path.isfile(path_file):
                os.remove(path_file)
        except Exception as e:
            if "O arquivo já está sendo usado por outro processo" in str(e):
                raise e

    @staticmethod
    def move_file(from_path_file, to_path_file):
        """
        Move a file from the source location to the destination location.

        :param from_path_file: Source file path.
        :type from_path_file: str
        :param to_path_file: Destination file path.
        :type to_path_file: str
        """
        shutil.copy2(from_path_file, to_path_file)

    @staticmethod
    def rename_file(from_path_file, to_path_file):
        """
        Rename a file from the source name to the destination name.

        :param from_path_file: Source file path.
        :type from_path_file: str
        :param to_path_file: Destination file path.
        :type to_path_file: str
        """
        from pathlib import Path
        Path(from_path_file).rename(to_path_file)

    @staticmethod
    def _ignore_unnamed_columns(df):
        """
        Remove columns with names starting with "unnamed" from a DataFrame.

        :param df: Input DataFrame.
        :type df: pandas.DataFrame
        :return: DataFrame with unnamed columns removed.
        :rtype: pandas.DataFrame
        """
        return df.select([col for col in df.columns if not str(col).lower().startswith('unnamed')])

    def _read_excel(self, file_path, **kwargs):
        """
        Read data from an Excel file.

        :param file_path: Path to the Excel file.
        :type file_path: str
        :param **kwargs: Additional keyword arguments passed to pandas.read_excel().
        :return: DataFrame containing the data from the Excel file.
        :rtype: pandas.DataFrame
        """
        # getting unnamend_columns value and removing it from kwargs
        unnamed_columns = kwargs.pop('unnamed_columns')
        try:
            df = pl.read_excel(file_path, **kwargs)
            if not unnamed_columns:
                df = self._ignore_unnamed_columns(df)
            if 'print' in kwargs:
                logging.info(df.head())
            return df
        except ValueError as e:
            logging.exception(e)

    def _read_csv(self, file_path, **kwargs):
        """
        Read data from a CSV file.

        :param file_path: Path to the CSV file.
        :type file_path: str
        :param **kwargs: Additional keyword arguments passed to pandas.read_csv().
        :return: DataFrame containing the data from the CSV file.
        :rtype: pandas.DataFrame
        """
        try:
            # getting unnamend_columns value and removing it from kwargs
            unnamed_columns = kwargs.pop('unnamed_columns')
            # Define default sep
            sep = ';' if 'sep' not in kwargs else kwargs.pop('sep')
            df = pl.read_csv(file_path, separator=sep, **kwargs)
            if not unnamed_columns:
                df = self._ignore_unnamed_columns(df)
            if 'print' in kwargs:
                logging.info(df.head())
            return df
        except ValueError as e:
            logging.exception(e)

    def _read_html(self, file_path, **kwargs):
        """
        Read data from an HTML file.

        :param file_path: Path to the HTML file.
        :type file_path: str
        :param **kwargs: Additional keyword arguments passed to pandas.read_html().
        :return: DataFrame containing the data from the HTML file.
        :rtype: pandas.DataFrame
        """
        try:
            # getting unnamend_columns and number file values and removing it form kwargs
            unnamed_columns = kwargs.pop('unnamed_columns')
            number_file = kwargs.pop('number_file')
            # Define default thousands
            thousands = '' if 'thousands' not in kwargs else kwargs.pop('thousands')

            # Define default decimal
            decimal = ',' if 'decimal' not in kwargs else kwargs.pop('decimal')

            # Define default flavor
            flavor = 'lxml' if 'flavor' not in kwargs else kwargs.pop('flavor')
            df = pd.read_html(file_path, thousands=thousands, decimal=decimal, flavor=flavor,
                              **kwargs)
            if 'print' in kwargs:
                [logging.info(d) for d in df]
            df = df[number_file]
            if not unnamed_columns:
                df = self._ignore_unnamed_columns(df)
            return df
        except ValueError as e:
            logging.exception(e)

    @staticmethod
    def _read_json(file_path):
        """
        Read data from a JSON file.

        :param file_path: Path to the JSON file.
        :type file_path: str
        :return: Dictionary containing the data from the JSON file.
        :rtype: dict
        """
        try:
            with open(file_path, "r") as readit:
                dict_file = json.load(readit)
            return dict_file
        except ValueError as e:
            logging.exception(e)

    @staticmethod
    def _read_bin(file_path):
        """
        Read binary data from a file and decode it.

        :param file_path: Path to the file.
        :type file_path: str
        :return: Decoded string representing the binary data.
        :rtype: str
        """
        try:
            with open(file_path, "rb") as f:
                bytes_f = f.read()
                valor = bytes_f.decode()
            return valor
        except ValueError as e:
            logging.exception(e)

    @staticmethod
    def _read_text(file_path):
        """
        Read text data from a file.

        :param file_path: Path to the file.
        :type file_path: str
        :return: Text data read from the file.
        :rtype: str
        """
        try:
            with open(file_path, "r") as f:
                read_file = f.read()
            return read_file
        except ValueError as e:
            logging.exception(e)

    @staticmethod
    def get_file_number_on_folder(folder, index_file):
        """
        Get the file at the specified index in a folder.

        :param folder: Path to the folder containing the files.
        :type folder: str
        :param index_file: Index of the file to retrieve.
        :type index_file: int
        :return: Full path to the file at the specified index.
        :rtype: str
        """
        files = os.listdir(folder)
        file_n = files[index_file]
        return os.path.join(folder, file_n)

    def import_file(self, file_path, **kwargs):
        """
        Import data from a file in various formats.

        :param file_path: (str) Path to the file to be imported.
        :param number_file: (int, optional) Index of the file to read if file_path is a directory. Default is 0.
        :param encoding: (str, optional) File encoding. Default is 'ISO-8859-1'.
        :param sep: (str, optional) Delimiter used in CSV files. Default is ';'.
        :param thousands: (str, optional) Thousands separator used in XLS files. Default is ''.
        :param decimal: (str, optional) Decimal separator used in XLS files. Default is ','.
        :param flavor: (str, optional) HTML parser flavor used in XLS files. Default is 'lxml'.
        :param unnamed_columns: (bool, optional) Whether to ignore unnamed columns in the imported data. Default is True.
        :param kwargs: Additional keyword arguments for specific file types.
        :return: (DataFrame, dict, str) Imported data, as a Pandas DataFrame, a dictionary or a string.

        This method imports data from a file in various formats, including Excel, CSV, XLS, JSON and text files. If the
        provided file_path is a directory, it reads the file at the specified index in the directory (default is the
        first file). The supported file formats are determined by the file extension, which is used to select the
        appropriate import function. Additional arguments can be passed as keyword arguments to the import functions,
        such as the delimiter used in CSV files or the encoding of the input file. By default, unnamed columns are
        ignored in the imported data.

        Example usage:
            >>> df = Handler.import_file('example.xlsx', sheet_name='Sheet1')
            >>> logging.info(df.head())
        """
        # Index file to read on the list os files on the path folder if file_path provide only the path folder
        number_file = 0 if 'number_file' not in kwargs else kwargs.pop('number_file')
        read_file = None

        # Verify if file_path provide the file with extention
        if len(file_path.split('.')) > 1:

            # configure unnamed_columns parameter if  keep or delete columns with the name unnamed in the name column
            unnamed_columns = False if 'unnamed_columns' not in kwargs else kwargs.pop('unnamed_columns')

            # Get the extention
            ext = file_path.split('.')[-1]

            # Define default encoding
            encoding = 'ISO-8859-1' if 'encoding' not in kwargs else kwargs.pop('encoding')

            # Read excel
            if 'xlsx' == ext.lower():
                read_file = self._read_excel(file_path, unnamed_columns=unnamed_columns, **kwargs)

            # Read csv
            elif 'csv' == ext.lower() or 'sswweb' == ext.lower():
                read_file = self._read_csv(file_path, unnamed_columns=unnamed_columns, encoding=encoding, **kwargs)

            # Read sheet html
            elif 'xls' == ext.lower():
                read_file = self._read_html(file_path, unnamed_columns=unnamed_columns, encoding=encoding,
                                            number_file=number_file, **kwargs)

            # Read json
            elif 'json' == ext.lower():
                read_file = self._read_json(file_path)

            # Read binary
            elif 'bin' == ext.lower():
                read_file = self._read_bin(file_path)

            # Try to read as text
            else:
                read_file = self._read_text(file_path)

            return read_file
        else:
            file_path = self.get_file_number_on_folder(file_path, number_file)
            return self.import_file(file_path, **kwargs)

    def to_csv(self, df, name_file, folder_to_salve, **kwargs):
        """
        Save a DataFrame as a CSV file.

        :param df: DataFrame to be saved.
        :type df: pandas.DataFrame
        :param name_file: Name of the output file.
        :type name_file: str
        :param folder_to_save: Folder to save the file.
        :type folder_to_save: str
        :param sep: Separator for the CSV file. Defaults to ';'.
        :type sep: str, optional
        :param index: Whether to include the DataFrame index in the file. Defaults to False.
        :type index: bool, optional
        """
        sep = ';' if 'sep' not in kwargs else kwargs.pop('sep')
        index = False if 'index' not in kwargs else kwargs.pop('index')
        self.create_folder(folder_to_salve)
        df.to_csv(f'{folder_to_salve}/{name_file}.csv', sep=sep, index=index, **kwargs)

    def to_excel(self, df, name_file, folder_to_salve, **kwargs):
        """
        Save a DataFrame as an Excel file.

        :param df: DataFrame to be saved.
        :type df: pandas.DataFrame
        :param name_file: Name of the output file.
        :type name_file: str
        :param folder_to_save: Folder to save the file.
        :type folder_to_save: str
        :param index: Whether to include the DataFrame index in the file. Defaults to False.
        :type index: bool, optional
        :param sheet_name: Name of the sheet in the Excel file. Defaults to the name_file parameter.
        :type sheet_name: str, optional
        """
        index = False if 'index' not in kwargs else kwargs.pop('index')
        sheet_name = name_file if 'sheet_name' not in kwargs else kwargs.pop('sheet_name')
        self.create_folder(folder_to_salve)
        df.to_excel(f'{folder_to_salve}/{name_file}.xlsx', sheet_name=sheet_name, index=index, **kwargs)

    @staticmethod
    def float_converter(table, column):
        """
        Convert the values in a specific column of a DataFrame to float.

        :param table: DataFrame containing the column.
        :type table: pandas.DataFrame
        :param column: Name of the column to be converted.
        :type column: str
        :return: The converted column.
        :rtype: pandas.Series
        """
        table[column] = table[column].astype('str')
        table[column] = list(
            map(lambda x: x.replace(' ', '').replace('.', '').replace(',', '.').strip(), table[column].tolist()))
        table[column] = list(
            map(lambda x: str(format(float(x), ".2f")), table[column].tolist()))
        return table[column]

    @staticmethod
    def adjust_encode_df_html(df, encoding: str = 'latin-1', deconding: str = 'utf-8'):
        """
            Adjusts the encoding of a DataFrame's columns and values.

            This function adjusts the encoding of a DataFrame's columns and values to the specified encoding and decoding.
            It returns the adjusted DataFrame.

            :param df: The DataFrame to adjust the encoding.
            :type df: pandas.DataFrame
            :param encoding: The encoding of the original DataFrame, defaults to 'latin-1'.
            :type encoding: str, optional
            :param decoding: The decoding to apply to the encoded DataFrame, defaults to 'utf-8'.
            :type decoding: str, optional
            :return: The adjusted DataFrame with the updated encoding.
            :rtype: pandas.DataFrame

            :process:
                1. Create a list of adjusted column names by encoding and decoding the original column names.
                2. Update the column names of the DataFrame with the adjusted column names.
                3. Iterate over each column in the DataFrame and encode and decode its values.
                4. Update the values in each column with the adjusted encoded and decoded values.
                5. Return the adjusted DataFrame.

            Example:
                >>> df = adjust_encode_df_html(df, encoding='latin-1', decoding='utf-8')
        """
        adjusted_columns = [col.encode(encoding).decode(deconding) for col in df.columns]
        df.columns = adjusted_columns

        for column in df.columns:
            df[column] = list(map(lambda x: str(x).encode(encoding).decode(deconding), df[column]))
        return df

    @staticmethod
    def convert_thousands_decimal_formats(table, column):
        """
        Convert values in a specific column of a DataFrame from a thousands and decimal format to a standard format.

        :param table: DataFrame containing the column.
        :type table: pandas.DataFrame
        :param column: Name of the column to be converted.
        :type column: str
        :return: The converted column.
        :rtype: pandas.Series
        """
        table[column] = table[column].astype('str')

        def adjust_decimal(valor):
            decimal_sep = [valor[i] for i in range(len(valor) - 1, -1, -1) if valor[i] == ',' or valor[i] == '.']
            if len(decimal_sep) > 0:
                if decimal_sep[0] == '.':
                    valor = valor.replace(',', '')
                    valor = valor.replace('.', ',')
                else:
                    valor = valor.replace('.', '')
            return valor

        table[column] = list(map(lambda x: adjust_decimal(x), table[column].tolist()))
        return table[column]

    def int_float_converter(self, table, column):
        """
        Convert values in a specific column of a DataFrame to either integers or floats, handling different formats.

        :param table: DataFrame containing the column.
        :type table: pandas.DataFrame
        :param column: Name of the column to be converted.
        :type column: str
        :return: The converted column.
        :rtype: pandas.Series
        """
        try:
            table[column] = table[column].astype('str')
            table[column] = self.remove_nan_from_table(table, column)
            table[column] = self.convert_thousands_decimal_formats(table, column)
            table[column] = list(
                map(lambda x: x.replace(' ', '').replace('#', '0').replace('*', '0').strip(), table[column].tolist()))

            lista_valores_int = list(
                map(lambda x: 'float' if ',' in str(x) and (int(str(x).split(',')[1]) > 0) else 'int',
                    table[column].tolist()))

            if 'float' not in lista_valores_int:
                table[column] = list(
                    map(lambda x: 0 if (len(x) <= 0) else int(float(str(x).replace(',', '.'))), table[column].tolist()))

            else:
                table[column] = list(
                    map(lambda x: x.replace(' ', '').replace('.', '').replace(',', '.').strip(),
                        table[column].tolist()))
                table[column] = list(map(lambda x: 0 if (len(x) <= 0) else float(x), table[column].tolist()))
                max_decimal_house = max([len(str(val).split('.')[1]) for val in table[column].tolist()])
                table[column] = list(map(lambda x: float(format(x, f".{max_decimal_house}f")), table[column].tolist()))
            return table[column]
        except ValueError:
            return table[column]

    @staticmethod
    def clear_invalid_characters_from_list(lista, replace: str = ' '):
        """
        Cleans invalid characters from a list of strings, replacing them with a specified character.

        :param lista: A list of strings to be cleaned.
        :param replace: The character to replace invalid characters with. Defaults to a space.
        :return: The cleaned list of strings.
        """
        lista = [re.sub(r'\W+', replace, valor).strip() if len(re.sub(r'\W+', replace, valor)) > 0 else valor.strip()
                 for valor in lista]
        return lista

    @staticmethod
    def remove_nan_from_table(table, column):
        """
        Remove NaN values from a specific column of a DataFrame.

        :param table: DataFrame containing the column.
        :type table: pandas.DataFrame
        :param column: Name of the column to remove NaN values from.
        :type column: str
        :return: The modified column.
        :rtype: pandas.Series
        """
        table[column] = list(map(lambda x: '' if (str(x) == 'nan') else x, table[column].tolist()))
        return table[column]

    @staticmethod
    def replace_none_for_zero_on_column(table, column):
        """
        Replace 'None' values with '0' in a specific column of a table.

        :param table: The table containing the column.
        :type table: pandas.DataFrame
        :param column: The name of the column to replace values in.
        :type column: str
        :return: The modified column with 'None' replaced by '0'.
        :rtype: pandas.Series
        """
        table[column] = list(map(lambda x: '0' if (str(x).lower() == 'none') else x, table[column].tolist()))
        return table[column]

    def clear_table(self, table):
        """
        Clear a DataFrame by removing NaN values and extra white spaces.

        :param table: DataFrame to be cleaned.
        :type table: pandas.DataFrame
        :return: The cleaned DataFrame.
        :rtype: pandas.DataFrame
        """
        table = table.fillna(value='')
        table = table.astype('str')
        for column in table.columns:
            # table[column] = self.remove_nan_from_table(table, column)
            table[column] = self.replace_none_for_zero_on_column(table, column)
            table[column] = list(map(lambda x: x.replace('  ', '').strip(), table[column].tolist()))
            table[column] = list(map(lambda x: '' if (x == 'NaT') else x, table[column].tolist()))
        return table

    @staticmethod
    def match_all_white_spaces(string: str):
        """
        Remove all white spaces from a string.

        :param string: The string to remove white spaces from.
        :type string: str
        :return: The modified string.
        :rtype: str
        """
        pattern = r"\s"
        matches = re.findall(pattern, string)
        for m in matches:
            string = string.replace(m, '')
        return string

    def replace_all_white_spaces(self, table, column):
        """
        Replace all white spaces in a specific column of a DataFrame.

        :param table: DataFrame containing the column.
        :type table: pandas.DataFrame
        :param column: Name of the column to replace white spaces in.
        :type column: str
        :return: The modified column.
        :rtype: pandas.Series
        """
        table[column] = list(map(lambda x: self.match_all_white_spaces(str(x)), table[column]))  # .tolist()
        return table[column]

    def identify_numeric_columns(self, table):
        """
        Identify numeric columns in a DataFrame.

        :param table: DataFrame to analyze.
        :type table: pandas.DataFrame
        :return: List of column names that contain only numeric values.
        :rtype: list[str]
        """
        table = table.copy()
        lista_columns_numericas = []
        for column in table.columns:

            table[column] = table[column].astype('str')
            table[column] = self.replace_all_white_spaces(table, column)
            table[column] = list(
                map(lambda x: x.lower().replace(' ', '').replace('.', '').replace(',', '')  # .replace('nan', '0')
                    .replace('#', '0').strip().replace('*', '0').strip() if len(str(x).strip()) > 0 else 0,
                    table[column]))  # .tolist()
            lista_cel = table[column].astype('str').str.isnumeric()
            lista_cel = list(
                map(lambda x: str(x), lista_cel))
            if 'False' not in lista_cel:
                lista_columns_numericas.append(column)

        return lista_columns_numericas

    @staticmethod
    def extract_file_from_rar(path_rar_file, path_to_save):
        from pyunpack import Archive
        # Extract the rar file
        Archive(path_rar_file).extractall(path_to_save)

    @staticmethod
    def extract_file_from_zip(path_zip_file, folder_to_save):
        """
        Extract a file from a ZIP archive.

        :param path_zip_file: Path to the ZIP file.
        :type path_zip_file: str
        :param folder_to_save: Folder to save the extracted file.
        :type folder_to_save: str
        """
        import zipfile
        # Open the zip file in read mode
        with zipfile.ZipFile(path_zip_file, 'r') as zip_ref:
            # Extract all the files
            zip_ref.extractall(folder_to_save)

    @staticmethod
    def zip_folder(folder_path, output_path):
        """
        Zip a folder.

        :param folder_path: Path of the folder to zip.
        :type folder_path: str
        :param output_path: Path of the output ZIP file.
        :type output_path: str
        """
        import zipfile
        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for root, dirs, files in os.walk(folder_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    zip_file.write(file_path, os.path.relpath(file_path, folder_path))

    def force_kill_process_by_pid(self, pid: int):
        """
        Forcefully kill a process by its PID on Windows.

        :param pid: Process ID of the process to kill.
        :type pid: int
        """
        os.system(f"taskkill /F /PID {pid}")

    def kill_process(self, key_word: str = 'main_bot', search_by: str = 'name'):
        """
        Kill a process by name.

        :param key_word: Name of the process to kill.
        :type key_word: str
        :param search_by: options 'name', 'exe', 'pid'.
        :type search_by: str
        """
        import psutil
        for proc in psutil.process_iter(['pid', 'name', 'exe']):  #
            logging.info([proc.info[key] for key in ['pid', 'name', 'exe']])
            try:
                if key_word.lower() in str(proc.info[search_by]).lower():
                    app_pid = proc.info['pid']
                    # Terminate the app process
                    app_process = psutil.Process(app_pid)
                    app_process.kill()
            except psutil.AccessDenied:
                logging.info(f"Access denied for process with PID {proc.info['pid']}. Continuing...")
                self.force_kill_process_by_pid(app_pid)

    @staticmethod
    def hash_string(s):
        """
        Compute the hash value of a string.

        :param s: Input string to hash.
        :type s: str
        :return: Hash value of the input string.
        :rtype: int
        """
        import hashlib
        return int(hashlib.sha256(s.encode('utf-8')).hexdigest(), 16)

    def sweep(self, path, file_to_seach, ignore_folders: list = []):
        """
        Recursively search for a file in a directory and its subdirectories.

        :param path: The path of the directory to start the search from.
        :type path: str
        :param file_to_search: The name of the file to search for.
        :type file_to_search: str
        :param ignore_folders: List of folders to ignore during the search (default: []).
        :type ignore_folders: list[str]
        :return: The path of the found file, or None if the file was not found.
        :rtype: str or None
        """
        if os.path.exists(path):
            try:
                contents = os.listdir(path)
            except Exception:
                return None

            files = [file for file in contents if '.' in str(file)]
            folders = [folder for folder in contents if '.' not in str(folder) and folder not in ignore_folders]
            if file_to_seach in ''.join(files):
                path_file = os.path.abspath(os.path.join(path, file_to_seach))
                return path_file
            else:
                for folder in folders:
                    path_child = os.path.join(path, folder)
                    resp = self.sweep(path_child, file_to_seach)

                    if resp is not None:
                        return resp

                return None
        else:
            return None

    def divide_table(self, df, qtde, name_file, folder_to_salve):
        """
        Divide a table into multiple parts and save each part as a CSV file.

        :param df: The DataFrame to be divided.
        :type df: pandas.DataFrame
        :param qtde: The number of parts to divide the table into.
        :type qtde: int
        :param name_file: The base name of the files to be saved.
        :type name_file: str
        :param folder_to_salve: The folder path to save the files.
        :type folder_to_salve: str
        """
        tamanho = len(df)
        partes = int(tamanho / qtde)
        for i in range(qtde):
            cut_ini = i * partes
            cut_fim = (i + 1) * partes - 1
            if i == qtde - 1:
                table_aux = df.loc[cut_ini:]
            else:
                table_aux = df.loc[cut_ini:cut_fim]
            filename = f'{name_file}_part_{i + 1}'
            self.to_csv(table_aux, filename, folder_to_salve)

    @staticmethod
    def get_str_format_datetime_by_switch_case(i):
        """
        Get the string format for datetime based on a switch case.

        :param i: The switch case value.
        :type i: int
        :return: The string format for datetime.
        :rtype: str
        """
        cases = {
            1: '%d/%m/%y',
            2: '%d/%m/%Y',
            3: '%m/%d/%y',
            4: '%m/%d/%Y',
            5: '%d/%m/%y %H:%M:%S',
            6: '%d/%m/%Y %H:%M:%S',
            7: '%m/%d/%Y %H:%M:%S',
            8: '%m/%d/%Y %H:%M:%S',
            9: '%d/%m/%y %H:%M',
            10: '%d/%m/%Y %H:%M',
            11: '%m/%d/%Y %H:%M',
            12: '%m/%d/%Y %H:%M'
        }
        return cases.get(i, 'Default Case')

    def _convert_column_to_datetime_with_strptime(self, table, column, format_d: str = "%d/%m/%y"):
        """
        Convert a column of a table to datetime using the strptime method.

        :param table: The table containing the column to convert.
        :type table: pandas.DataFrame
        :param column: The name of the column to convert.
        :type column: str
        :param format_d: The initial datetime format to use (default: "%d/%m/%y").
        :type format_d: str
        :return: The converted column.
        :rtype: pandas.Series
        """
        for i in range(12):
            try:
                table[column] = table[column].apply(
                    (lambda x: datetime.strptime(x, format_d) if (x != '') else 'NULL'))
                table[column] = table[column].str.to_datetime(format=format_d)
                table[column] = list(
                    map(lambda x: x if (x is not None) and len(str(x)) >= 8 else 'NULL', table[column].tolist()))
                break
            except Exception as e:
                logging.info(e)
                format_d = self.get_str_format_datetime_by_switch_case(i + 1)
                logging.info(format_d)
        return table[column]

    def _convert_column_to_datetime_with_datetime64(self, table, column, format_d: str = "%d/%m/%y"):
        """
        Convert a column of a table to datetime using the datetime64 data type.

        :param table: The table containing the column to convert.
        :type table: pandas.DataFrame
        :param column: The name of the column to convert.
        :type column: str
        :param format_d: The initial datetime format to use (default: "%d/%m/%y").
        :type format_d: str
        :return: The converted column.
        :rtype: pandas.Series
        """
        for i in range(12):
            try:
                table[column] = table[column].astype('datetime64[ns]')
                table[column] = table[column].str.to_datetime(format=format_d)
                table[column] = list(map(lambda x: x if (len(str(x)) >= 10) else 'NULL',
                                         table[column].tolist()))
                break
            except Exception as e:
                logging.info(e)
                format_d = self.get_str_format_datetime_by_switch_case(i + 1)
                logging.info(format_d)
        return table[column]

    def convert_column_to_datetime(self, table, column):
        """
        Convert a column of a table to datetime.

        The method tries different approaches to convert the column to datetime, including using strptime and datetime64.

        :param table: The table containing the column to convert.
        :type table: pandas.DataFrame
        :param column: The name of the column to convert.
        :type column: str
        :return: The converted column.
        :rtype: pandas.Series
        """

        # identifying if the first value seem like date value if yes get only the first 10 characters
        # primeiro = ''
        # for valor in table[column]:
        #     if str(valor) != 'nan' and len(str(valor)) > 0:
        #         primeiro = str(valor)
        # if '-' in primeiro:
        #     table[column] = list(map(lambda x: str(x)[:10], table[column].tolist()))

        # Removing white spaces and nan values
        # table[column] = list(map(lambda x: ))
        table[column] = self.remove_nan_from_table(table, column)

        # Trying many ways to convert the column to datetime
        try:
            table[column] = self._convert_column_to_datetime_with_strptime(table, column)
        except Exception:
            try:
                table[column] = table[column].str.to_datetime(format='%d/%m/%y')
            except Exception:
                try:
                    table[column] = self._convert_column_to_datetime_with_datetime64(table, column)
                except Exception as e:
                    logging.exception(e)

        # replacing empty values with 'NULL' value to be set on the database
        table[column] = list(
            map(lambda x: x if len(str(x)) >= 8 else 'NULL', table[column].tolist()))

        return table[column]

    @staticmethod
    def get_str_format_time_by_switch_case(i):
        """
        Get the time format based on the switch case value.

        :param i: The switch case value.
        :type i: int
        :return: The corresponding time format.
        :rtype: str
        """
        cases = {
            1: '%H:%M:%S',
            2: '%H:%M',
        }
        return cases.get(i, 'Default Case')

    def convert_column_to_time_with_strptime(self, table, column, format_t: str = '%H:%M:%S'):
        """
        Convert a column of a table to time using the strptime function.

        :param table: The table containing the column to convert.
        :type table: pandas.DataFrame
        :param column: The name of the column to convert.
        :type column: str
        :param format_t: The format string for time conversion, defaults to '%H:%M:%S'.
        :type format_t: str, optional
        :return: The converted column.
        :rtype: pandas.Series
        """
        for i in range(2):
            try:
                table[column] = table[column].apply(
                    (lambda x: datetime.strptime(str(x).strip(), format_t) if (x != '') else None))
                table[column] = table[column].str.to_datetime(format=format_t)
                break
            except Exception as e:
                logging.info(e)
                format_t = self.get_str_format_time_by_switch_case(i + 1)
        # replacing empty values with 'NULL' value to be set on the database
        table[column] = list(
            map(lambda x: x if len(str(x)) >= 4 else 'NULL', table[column].tolist()))
        return table[column]

    def _convert_column_to_time_with_datetime64(self, table, column, format_t: str = '%H:%M:%S'):
        """
        Convert a column of a table to time using the datetime64 parameter.

        :param table: The table containing the column to convert.
        :type table: pandas.DataFrame
        :param column: The name of the column to convert.
        :type column: str
        :param format_t: The format string for time conversion, defaults to '%H:%M:%S'.
        :type format_t: str, optional
        :return: The converted column.
        :rtype: pandas.Series
        """
        for i in range(2):
            try:
                table[column] = table[column].str.to_datetime(format=format_t)
                table[column] = table[column].astype('datetime64[ns]')
                table[column] = table[column].str.to_datetime(format=format_t).dt.time
                break
            except Exception as e:
                logging.info(e)
                format_t = self.get_str_format_time_by_switch_case(i + 1)

        # replacing empty values with 'NULL' value to be set on the database
        table[column] = list(
            map(lambda x: x if len(str(x)) >= 4 else 'NULL', table[column].tolist()))
        return table[column]

    def _convert_column_to_time(self, table, column):
        """
        Convert a column of a table to time.

        The method tries different approaches to convert the column to time, including using strptime and datetime64.

        :param table: The table containing the column to convert.
        :type table: pandas.DataFrame
        :param column: The name of the column to convert.
        :type column: str
        :return: The converted column.
        :rtype: pandas.Series
        """
        # try many ways to convert the column to time format
        try:
            # converting with strptime function to format H:M:S
            table[column] = self.convert_column_to_time_with_strptime(table, column)
        except Exception as e:
            logging.exception(e)
            try:
                # converting with datetime64 parameter to format H:M
                table[column] = self._convert_column_to_time_with_datetime64(table, column)
            except Exception as e:
                logging.exception(e)

        return table[column]

    def convert_table_columns_to_datetime(self, table, ignore_type_columns: list = None, dtypes: dict = None):
        """
        Convert columns of a table to datetime based on column names and provided types.

        This method iterates through the columns of the table and converts columns to datetime if their names
        match the specified types.

        :param table: The table to convert.
        :type table: pandas.DataFrame
        :param ignore_type: A list of types to ignore during conversion, defaults to None.
        :type ignore_type: list, optional
        :param dtypes: A dictionary of types to match column names against, defaults to None.
        :type dtypes: dict, optional
        :return: The converted table.
        :rtype: pandas.DataFrame
        """
        for column in table.columns:
            column = str(column)
            # logging.info(column)
            # verify if name column seem like a name column date
            if ignore_type_columns is None or column.lower() not in ignore_type_columns:
                if True in [True for arg in dtypes['datetime'] if arg.lower() in column.lower()]:
                    table[column] = self.convert_column_to_datetime(table, column)

            # verify if name column seem like a name column time
            if ignore_type_columns is None or column.lower() not in ignore_type_columns:
                if True in [True for arg in dtypes['time'] if arg.lower() in column.lower()]:
                    table[column] = self._convert_column_to_time(table, column)
            # logging.info(table[column].info())
        return table

    def convert_table_types(self, table, ignore_type_columns: list = None, dtypes: dict = None):
        """
        Convert columns of a table to their specified types.

        This method converts columns of the table to their specified types, including datetime and numeric types.

        :param table: The table to convert.
        :type table: pandas.DataFrame
        :param ignore_type: A list of types to ignore during conversion, defaults to None.
        :type ignore_type: list, optional
        :param dtypes: A dictionary specifying a list of match words for 'datetime' columns and 'time' columns, defaults
         if None is {'datetime':['data','dia'],'time':['hora']}.
        :type dtypes: dict, optional
        :return: The converted table.
        :rtype: pandas.DataFrame
        """
        if dtypes is None:
            dtypes = {}
        if ignore_type_columns is not None:
            ignore_type_columns = [col.lower() for col in ignore_type_columns]

        dtypes['datetime'] = ['data', 'dia'] if 'datetime' not in dtypes.keys() else dtypes['datetime']
        dtypes['time'] = ['hora'] if 'time' not in dtypes.keys() else dtypes['time']

        # try to convert columns to datetime
        table = self.convert_table_columns_to_datetime(table, ignore_type_columns, dtypes)

        # try to identify and convert numeric columns
        numeric_columns = self.identify_numeric_columns(table)
        for column_n in numeric_columns:
            if ignore_type_columns is None or column_n.lower() not in ignore_type_columns:
                table[column_n] = self.int_float_converter(table, column_n)
        return table

    @staticmethod
    def get_date_back(days):
        return datetime.now() - timedelta(days)

    @staticmethod
    def get_date_forward(days):
        return datetime.now() + timedelta(days)

    @staticmethod
    def _verify_month_range(month):
        # verify month range
        if int(month) not in range(1, 13):
            raise ValueError('Expected month value in range 1-12')

    @staticmethod
    def _verify_year_format(year):
        # Verrify year format
        if len(year) == 2:
            date_format = '%d/%m/%y'
        elif len(year) == 4:
            date_format = '%d/%m/%Y'
        else:
            raise ValueError("Expected year format 'yy' or 'yyyy'")
        return date_format

    @staticmethod
    def _verify_numeric_month_year(month, year):
        # verifying if the params are numeric
        if not str(month).isnumeric() or not str(year).isnumeric():
            raise ValueError('Numeric values expected!')

    def get_full_range_days_from_month(self, month: str or int = 0, year: str or int = 0, format_date: str = ''):
        self._verify_numeric_month_year(month, year)

        # get the month before
        month_before = str(date.today().month - 1) if str(date.today().month - 1) != '0' else '12'

        # if month is zero define as month before
        month = str(month) if str(month) != '0' else month_before

        # adjust the month length
        month = '0' + str(month) if len(month) < 2 else month

        # verify month range
        self._verify_month_range(month)

        # define next month
        month_prox = str(int(month) + 1) if str(int(month) + 1) != '13' else '01'
        month_prox = '0' + str(month_prox) if len(month_prox) < 2 else month_prox

        # if year is zero define current year
        year = str(year) if str(year) != '0' else str(date.today().year)

        # Verrify year format
        date_format = self._verify_year_format(year)

        # get the firt day month in datetime
        first_day_month = datetime.strptime(f'01/{month}/{year}', date_format)

        # verify if month_prox will lead to next year or not
        year = year if month_prox != '01' else str(int(year) + 1)
        year = '0' + str(year) if len(year) < 2 else year

        # get the firt day from next month
        first_day_prox_month = datetime.strptime(f'01/{month_prox}/{year}', date_format)

        # get the last day month
        last_day_month = first_day_prox_month - timedelta(1)

        if len(str(format_date)) > 0:
            last_day_month = self.string_format_br_date(last_day_month, format_date)
            first_day_month = self.string_format_br_date(first_day_month, format_date)

        return first_day_month, last_day_month

    def get_range_days_from_current_month(self, end_date_limit: int = 0, format_date: str = ''):
        current_date = self.get_date_back(end_date_limit)
        month = '0' + str(current_date.month) if len(str(current_date.month)) < 2 else current_date.month
        first_date = datetime.strptime(f'01/{month}/{current_date.year}', '%d/%m/%Y')
        if len(str(format_date)) > 0:
            current_date = self.string_format_br_date(current_date, format_date)
            first_date = self.string_format_br_date(first_date, format_date)
        return first_date, current_date

    @staticmethod
    def convert_string_date_to_datetime(date_string, date_format: str = "%d/%m/%Y"):
        # Convert the date string to a datetime object
        return datetime.strptime(date_string, date_format)

    @staticmethod
    def string_format_br_date(date: datetime, format_date: str = '6'):
        """
        This function takes a datetime object and a format code as inputs and returns the date formatted in various Brazilian formats.

        :param date: The input date as a datetime object
        :type date: datetime

        :param format_date: Format code to specify the desired date format.
                            Available options are:
                             - '6' : ddmmyy
                             - '8a': ddmmyyyy
                             - '8b': dd/mm/yy
                             - '8c': dd-mm-yy
                             - '10a': dd/mm/yyyy
                             - '10b': dd-mm-yyyy
        :type format_date: str

        :return: The date formatted in the specified Brazilian date format
        :rtype: str

        :process:
            1. Extract the day and month as two-digit strings.
            2. Extract the year in two formats: 'yy' and 'yyyy'.
            3. Format the date string based on the format code provided.

        Example:
            >>> from datetime import datetime
            >>> string_format_br_date(datetime(2023, 1, 1), '8a')
            '01012023'
        """
        # adjusting day and month length
        day = '0' + str(date.day) if len(str(date.day)) < 2 else str(date.day)
        month = '0' + str(date.month) if len(str(date.month)) < 2 else str(date.month)

        # defining the 'yy' and 'yyyy' year formats
        year_2 = str(date.year)[-2:]
        year_4 = str(date.year)

        # formating according the desired br format
        # 010123
        if '6' in str(format_date):
            return f'{day}{month}{year_2}'
        # 01012023
        if '8a' in str(format_date):
            return f'{day}{month}{year_4}'
        # 01/01/23
        if '8b' in str(format_date):
            return f'{day}/{month}/{year_2}'
        # 01-01-23
        if '8c' in str(format_date):
            return f'{day}-{month}-{year_2}'
        # 01/01/2023
        if '10a' in str(format_date):
            return f'{day}/{month}/{year_4}'
        # 01-01-2023
        if '10b' in str(format_date):
            return f'{day}-{month}-{year_4}'
        if format_date is None:
            return date

    @staticmethod
    def add_left_zero(value, len_desired):
        zeros = ''.join(['0' for _ in range(len_desired - len(str(value)))])
        return zeros + str(value)

    @staticmethod
    def get_day_greetings():
        current_hour = datetime.now().hour
        if current_hour < 12:
            greeting = 'Bom dia! Como sr(a) está?'
        elif 12 <= current_hour < 18:
            greeting = 'Boa tarde! Como sr(a) está?'
        else:
            greeting = 'Boa noite! Como sr(a) está?'
        return greeting

    @staticmethod
    def match_numeric(value):
        return ''.join(re.findall('\d', str(value)))

    @staticmethod
    def match_letter(value):
        return ''.join(re.findall(r'[a-zA-Z]+', value))

    @staticmethod
    def change_letters_per_number(text):
        import string
        # Get the letters of alphabet
        alphabet = str(string.ascii_lowercase)

        # Convert the text to lower case to match with the lower letters in the alphabet
        text = text.lower()

        # Replace letter by the position number of this letter in the alphabet and keep char number
        number_string = int(''.join(str(alphabet.find(c) + 1) if len(re.findall('\d', c)) == 0 else c for c in text))

        return number_string

    @staticmethod
    def convert_img_to_pdf(path_img):
        from PIL import Image
        base_path = path_img.split('.')[0]
        pdf_path = f'{base_path}.pdf'

        # Open an image file
        with Image.open(path_img) as image:
            # Convert the image to RGB, even if it's already in RGB mode, to ensure the image is saved as a color image
            imagelist = [image.convert('RGB')]
            image.save(pdf_path, save_all=True, append_images=imagelist)

    @staticmethod
    def partial_comparation(list_1, list_2, confidence: int = 70):
        from fuzzywuzzy import fuzz

        def is_similar(str1, str2):
            return fuzz.ratio(str1, str2) >= confidence

        result = []
        for word1 in list_1:
            found = False
            for word2 in list_2:
                if is_similar(word1, word2):
                    found = True
                    break
            result.append(found)

        return result

    @staticmethod
    def find_wrong_rows_csv_file(file_path, sep: str = ';'):
        with open(file_path, 'r') as f:
            read_ = f.read()
            rows = read_.split('\n')
            len_columns = [len(row.split(sep)) for row in rows]
            filt_len_columns = [len_ for len_ in len_columns if len_ > 1]
            val_max = max(filt_len_columns)
            val_min = min(filt_len_columns)
            wrong_rows = [i for i, row in enumerate(rows) if len(row.split(sep)) != val_min]
            logging.info(val_max, val_min, wrong_rows)
            return wrong_rows

    @staticmethod
    def color_table_line(file_path, color_argb_background: str = 'FFFFFF00', color_argb_values: str = None,
                         index_line=1):
        """
            Colorizes a specific line in a table in an Excel file.

            This function applies a background color and/or a font color to a specific line in a table in an Excel file.
            The file is modified in-place.

            :param file_path: The path to the Excel file.
            :type file_path: str
            :param color_argb_background: The ARGB value of the background color, defaults to 'FFFFFF00'.
            :type color_argb_background: str, optional
            :param color_argb_values: The ARGB value of the font color, defaults to None.
            :type color_argb_values: str, optional
            :param index_line: The index of the line to colorize, defaults to 1.
                               If multiple indices are provided, each line will be colorized.
            :type index_line: int or list[int], optional

            :process:
                1. Load the existing workbook from the file.
                2. Select the active worksheet.
                3. If index_line is an integer, convert it to a list containing a single element.
                4. Iterate over each line in index_line.
                5. Select the row corresponding to the line.
                6. If color_argb_values is not None, create a font with the specified color.
                   Apply this font to all cells in the row.
                7. If color_argb_background is not None, create a fill pattern with the specified background color.
                   Apply this fill pattern to all cells in the row.
                8. Save the modified workbook back to the file.

            Example:
                >>> file_path = "table.xlsx"
                >>> color_table_line(file_path, color_argb_background='FFFFFF00', color_argb_values=None, index_line=1)
        """
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Color, PatternFill

        # Load an existing workbook
        wb = load_workbook(file_path)
        ws = wb.active

        if isinstance(index_line, int):
            index_line = [index_line]

        for line in index_line:
            # Select the first row
            header_row = ws[line]

            # color de values
            if color_argb_values is not None:
                font = Font(color=Color(color_argb_values))

                # Apply this font to all cells in the header row
                for cell in header_row:
                    cell.font = font

            # color de background
            if color_argb_background is not None:
                yellow_fill = PatternFill(start_color=color_argb_background, end_color=color_argb_background,
                                          fill_type="solid")

                # Apply this font to all cells in the header row
                for cell in header_row:
                    cell.fill = yellow_fill

        # Save the workbook
        wb.save(file_path)

    def get_list_start_end_dates_by_range_days_on_year(self, start_date: str = '', end_date: str = '', step: int = 15,
                                                       year: int = 2023, format_date='6'):
        ini_ = f'01/01/{year}' if len(start_date) == 0 else start_date
        ini_ = self.convert_string_date_to_datetime(ini_)
        end_ = self.get_date_back(0) if len(end_date) == 0 else self.convert_string_date_to_datetime(end_date)
        diff = (end_ - ini_).days
        range_ = int(diff / step)
        list_start_dates = [ini_]
        list_end_dates = []
        next_date_ = ini_
        for i in range(range_):
            last_date = next_date_ + timedelta(1)
            next_date_ = next_date_ + timedelta(step)
            if i == range_ - 1:
                next_date_ = end_
            list_end_dates.append(next_date_)
            if i > 0:
                list_start_dates.append(last_date)
        for i in range(range_):
            list_start_dates[i] = self.string_format_br_date(list_start_dates[i], format_date=format_date)
            list_end_dates[i] = self.string_format_br_date(list_end_dates[i], format_date=format_date)
        return list_start_dates, list_end_dates

    def append_text_to_file(self, text: str = '', name_file: str = '', path_to_save: str = '',
                            create_if_not_exist: bool = False):
        """
        Creates a .txt file at a specified location with provided text. If the file already exists,
        the method can either overwrite it or read its content based on the 'subs' argument.
        The method first ensures that the path to save the file exists, creating it if it doesn't.

        :param text: The text to write into the file. Default is an empty string.
        :type text: str

        :param name_file: The name of the file to be created. Default is an empty string.
        :type name_file: str

        :param path_to_save: The directory where the file should be saved. Default is the current working directory.
        :type path_to_save: str

        :param create_if_not_exist: create the file if not exists. Default is False.
        :type create_if_not_exist: bool

        :return: The text written to the file or read from it.
        :rtype: str

        :Example:

            self._create_file_txt("Hello, world!", "test", "/path/to/save", True) # creates/overwrites a file named "test.txt" with the text "Hello, world!" at "/path/to/save"

        """
        path_to_save = self.create_folder(path_to_save)
        full_path_file = f'{path_to_save}\{name_file}.txt'

        if os.path.exists(full_path_file) or create_if_not_exist:
            with open(full_path_file, 'a') as f:
                f.write(text)
        return text

    @staticmethod
    def get_day_of_week(isoweekday):
        switcher = {
            1: "SEGUNDA",
            2: "TERCA",
            3: "QUARTA",
            4: "QUINTA",
            5: "SEXTA",
            6: "SABADO",
            7: "DOMINGO",
        }

        return switcher.get(isoweekday, isoweekday)

    @staticmethod
    def thread_it(qtd_threads, list_params, function):
        from tqdm import tqdm
        import threading as td
        threads = [td.Thread() for _ in range(qtd_threads)]
        counter = 0
        for param in tqdm(list_params, desc="Threading", unit="item"):
            threads[counter] = td.Thread(target=lambda: function(param))
            threads[counter].start()
            counter += 1
            if counter >= qtd_threads:
                [threads[i].join() for i in range(counter)]
                counter = 0
        if counter < qtd_threads:
            [threads[i].join() for i in range(counter)]

    @staticmethod
    def scan_image(image_path, output_path):
        import cv2
        # Read the image
        image = cv2.imread(image_path, cv2.IMREAD_COLOR)

        # Convert the image to grayscale
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        height, width = gray.shape
        gray = cv2.resize(gray, (width * 4, height * 4), interpolation=cv2.INTER_LINEAR)

        processed = cv2.equalizeHist(gray)

        processed = cv2.GaussianBlur(processed, (3, 3), 0)

        # Dilation can be used to strengthen the edges
        kernel = np.ones((3, 3), np.uint8)
        dilated = cv2.dilate(processed, kernel, iterations=1)

        # Combine the dilated edges with the original image
        processed = cv2.addWeighted(processed, 0.7, dilated, 0.8, 0.2)

        # Save the preprocessed image
        cv2.imwrite(output_path, processed)

    @staticmethod
    def remove_background(path_from, path_to):
        from rembg import remove
        from PIL import Image
        # Processing the image
        input_img = Image.open(path_from)

        # Removing the background from the given Image
        output = remove(input_img)

        # Saving the image in the given path
        output.save(path_to)

    @staticmethod
    def is_iterable(obj):
        if isinstance(obj, str) or isinstance(obj, dict):
            return False  # Treat strings as non-iterable
        try:
            iter(obj)
            return True
        except TypeError:
            return False

    def _append_results(self, report, result):
        if self.is_iterable(result):
            for k, r in enumerate(result):
                try:
                    report[f'result_{k + 1}'] += [r]
                except Exception as e:
                    logging.exception(e)
                    report[f'result_{k + 1}'] = [r]
        else:
            try:
                report[f'result'] += [result]
            except Exception as e:
                logging.exception(e)
                report[f'result'] = [result]

    def get_report_from_interaction(self, func, list_interaction):
        report = {'interactor': list_interaction}
        for param in list_interaction:
            try:
                result = func(param)
                self._append_results(report, result)
            except Exception as e:
                logging.exception(e)
                self._append_results(report, [f'ERROR: {str(e)}' for _ in list(report.keys())[1:]])
        # logging.info(json.dumps(report,indent=6))
        logging.info(report)
        report = pl.DataFrame(report)
        return report


if __name__ == '__main__':
    hd = Handler()
    hd.append_text_to_file('teste', 'teste', 'teste_append', create_if_not_exist=True)
